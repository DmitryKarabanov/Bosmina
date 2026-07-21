import re
import math
import base64
import json
from collections import defaultdict
from pyvis.network import Network
from functools import lru_cache
import networkx as nx

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install it: pip install openpyxl")
    exit(1)

# ==========================================
# 1. SETTINGS
# ==========================================
NEXUS_FILE = "bosmina_popart.nex"
XLSX_FILE = "_resources.xlsx"
OUTPUT_HTML = "Bosmina_TCS_MedianJoining.html"
TCS_CONFIDENCE = 0.90
MJ_EPSILON = 1.0
MAX_MEDIANS = 48
MAX_MJ_ITERATIONS = 1000

IUPAC = {
    'A': {'A'}, 'C': {'C'}, 'G': {'G'}, 'T': {'T'}, 'U': {'T'},
    'R': {'A', 'G'}, 'Y': {'C', 'T'}, 'S': {'G', 'C'}, 'W': {'A', 'T'},
    'K': {'G', 'T'}, 'M': {'A', 'C'}, 'B': {'C', 'G', 'T'}, 'D': {'A', 'G', 'T'},
    'H': {'A', 'C', 'T'}, 'V': {'A', 'C', 'G'}, 'N': {'A', 'C', 'G', 'T'},
    '-': set(), '?': set(), '.': set(), '*': set()
}

print("Reading NEXUS file...")
with open(NEXUS_FILE, 'r', encoding='utf-8') as f:
    content = f.read()

# ==========================================
# 2. NEXUS PARSING
# ==========================================
matrix_match = re.search(r'BEGIN DATA;.*?MATRIX\s*(.*?)\s*;\s*END;', content, re.DOTALL | re.IGNORECASE)
seqs = {}
for line in matrix_match.group(1).strip().split('\n'):
    parts = line.strip().split()
    if len(parts) >= 2:
        seqs[parts[0]] = parts[1]

traits_match = re.search(r'BEGIN TRAITS;.*?Matrix\s*(.*?)\s*;\s*END;', content, re.DOTALL | re.IGNORECASE)
trait_labels_match = re.search(r'TraitLabels\s+(.*?);', content, re.IGNORECASE)
trait_labels = trait_labels_match.group(1).strip().split()

seq_traits = {}
for line in traits_match.group(1).strip().split('\n'):
    parts = line.strip().split()
    if len(parts) >= 2:
        seq_name = parts[0]
        counts = [int(x) for x in parts[1].split(',')]
        seq_traits[seq_name] = dict(zip(trait_labels, counts))

print(f"Found {len(seqs)} sequences.")

# ==========================================
# 2.5. READ HAPLOGROUP DATA FROM XLSX
# ==========================================
print(f"Reading haplogroup data from {XLSX_FILE}...")
wb = openpyxl.load_workbook(XLSX_FILE, read_only=True)
ws = wb.active

header = [cell.value for cell in ws[1]]
seq_col_idx = None
hg_col_idx = None
for i, h in enumerate(header):
    if h and str(h).strip().lower() == 'seq':
        seq_col_idx = i
    if h and str(h).strip().lower() == 'haplogroup':
        hg_col_idx = i

if seq_col_idx is None or hg_col_idx is None:
    print(f"ERROR: Could not find 'seq' or 'Haplogroup' columns in {XLSX_FILE}")
    print(f"Found columns: {header}")
    exit(1)

seq_to_haplogroup = {}
for row in ws.iter_rows(min_row=2):
    seq_name = row[seq_col_idx].value
    hg = row[hg_col_idx].value
    if seq_name and hg:
        seq_to_haplogroup[str(seq_name).strip()] = str(hg).strip()

wb.close()
print(f"Loaded {len(seq_to_haplogroup)} haplogroup assignments.")

# Проверяем совпадение имён
nexus_names = set(seqs.keys())
xlsx_names = set(seq_to_haplogroup.keys())
missing = nexus_names - xlsx_names
if missing:
    print(f"WARNING: {len(missing)} NEXUS sequences not found in xlsx (first 5): {list(missing)[:5]}")

# ==========================================
# 3. GENETICALLY CORRECT DISTANCE CALCULATION
# ==========================================
def calc_dna_dist(seq1, seq2):
    dist = 0
    min_len = min(len(seq1), len(seq2))
    for i in range(min_len):
        c1, c2 = seq1[i].upper(), seq2[i].upper()
        set1 = IUPAC.get(c1, set())
        set2 = IUPAC.get(c2, set())
        if not set1 or not set2:
            continue
        if set1.isdisjoint(set2):
            dist += 1
    return dist

# ==========================================
# 4. COLLAPSING INTO UNIQUE HAPLOTYPES
# ==========================================
haplotypes = defaultdict(list)
for name, seq in seqs.items():
    haplotypes[seq].append(name)

unique_haps = list(haplotypes.keys())
hap_names = {seq: f"Hap_{i+1:03d}" for i, seq in enumerate(unique_haps)}
print(f"Collapsed to {len(unique_haps)} unique haplotypes.")

# Определяем гаплогруппу для каждого гаплотипа (мажоритарная)
hap_to_haplogroup = {}
for seq, names in haplotypes.items():
    hg_counts = defaultdict(int)
    for name in names:
        if name in seq_to_haplogroup:
            hg_counts[seq_to_haplogroup[name]] += 1
    if hg_counts:
        hap_to_haplogroup[seq] = max(hg_counts, key=hg_counts.get)

# Палитра для гаплогрупп
HG_PALETTE = [
    '#e6194b', '#3cb44b', '#4363d8', '#f58231', '#911eb4',
    '#42d4f4', '#f032e6', '#bfef45', '#fabed4', '#469990',
    '#dcbeff', '#9A6324', '#800000', '#aaffc3', '#808000',
    '#ffd8b1', '#000075', '#a9a9a9', '#ffe119', '#008080'
]

unique_haplogroups = sorted(set(hap_to_haplogroup.values()))
hg_colors = {hg: HG_PALETTE[i % len(HG_PALETTE)] for i, hg in enumerate(unique_haplogroups)}
print(f"Found {len(unique_haplogroups)} haplogroups: {unique_haplogroups}")

# ==========================================
# FAST DISTANCE CALCULATION WITH CACHING
# ==========================================
@lru_cache(maxsize=None)
def calc_dna_dist_cached(seq1, seq2):
    if seq1 > seq2:
        seq1, seq2 = seq2, seq1
    dist = 0
    min_len = min(len(seq1), len(seq2))
    for i in range(min_len):
        c1, c2 = seq1[i].upper(), seq2[i].upper()
        set1 = IUPAC.get(c1, set())
        set2 = IUPAC.get(c2, set())
        if not set1 or not set2:
            continue
        if set1.isdisjoint(set2):
            dist += 1
    return dist

# ==========================================
# 5. TCS THRESHOLD CALCULATION
# ==========================================
print("Calculating probabilistic TCS threshold...")

def poisson_pmf(k, lam):
    try:
        return (lam**k) * math.exp(-lam) / math.factorial(k)
    except (OverflowError, ValueError):
        return 0.0

def tcs_connection_limit(n_haplotypes, confidence=0.95):
    if n_haplotypes <= 1:
        return 1
    lam = 2.0 * math.log(n_haplotypes)
    cumul = 0.0
    for k in range(1, 1000):
        cumul += poisson_pmf(k, lam)
        if cumul >= (1.0 - confidence):
            return k
    return 1

TCS_LIMIT = tcs_connection_limit(len(unique_haps), TCS_CONFIDENCE)
print(f"Automatic TCS threshold: <= {TCS_LIMIT} mutations.")

# ==========================================
# 6. MEDIAN-JOINING NETWORK
# ==========================================
print("Building Minimum Spanning Network...")

n = len(unique_haps)
dist_matrix = [[0] * n for _ in range(n)]
for i in range(n):
    for j in range(i + 1, n):
        d = calc_dna_dist(unique_haps[i], unique_haps[j])
        dist_matrix[i][j] = d
        dist_matrix[j][i] = d

def prim_mst(n_nodes, dist_mat):
    in_mst = [False] * n_nodes
    in_mst[0] = True
    edges = []
    for _ in range(n_nodes - 1):
        min_w = float('inf')
        best_edge = None
        for u in range(n_nodes):
            if not in_mst[u]: continue
            for v in range(n_nodes):
                if in_mst[v]: continue
                if dist_mat[u][v] < min_w:
                    min_w = dist_mat[u][v]
                    best_edge = (u, v, min_w)
        if best_edge:
            u, v, w = best_edge
            edges.append((u, v, w))
            in_mst[v] = True
    return edges

mst_edges = prim_mst(n, dist_matrix)
print(f"MST built: {len(mst_edges)} edges.")

adj = defaultdict(set)
for u, v, w in mst_edges:
    adj[u].add(v)
    adj[v].add(u)

def compute_single_median(s1, s2, s3):
    if not (len(s1) == len(s2) == len(s3)):
        return None
    median_chars = []
    for i in range(len(s1)):
        c1, c2, c3 = s1[i], s2[i], s3[i]
        if c1 == c2 or c1 == c3:
            median_chars.append(c1)
        elif c2 == c3:
            median_chars.append(c2)
        else:
            median_chars.append(c1)
    return ''.join(median_chars)

print("Searching for median vectors...")
all_sequences = set(unique_haps)
all_seq_list = list(unique_haps)
median_flags = [False] * len(unique_haps)
iteration = 0
changed = True
medians_added = 0

def find_triplets(adj_dict):
    triplets = set()
    for u in adj_dict:
        neighbors_u = list(adj_dict[u])
        for i in range(len(neighbors_u)):
            for j in range(i+1, len(neighbors_u)):
                v, w = neighbors_u[i], neighbors_u[j]
                triplets.add(tuple(sorted([u, v, w])))
    return list(triplets)

while changed and iteration < MAX_MJ_ITERATIONS and medians_added < MAX_MEDIANS:
    iteration += 1
    changed = False
    triplets = find_triplets(adj)
    print(f"  Iteration {iteration}: analyzing {len(triplets)} triplets...")

    candidates = []
    for triplet in triplets:
        u, v, w = triplet
        s1, s2, s3 = all_seq_list[u], all_seq_list[v], all_seq_list[w]
        
        median_seq = compute_single_median(s1, s2, s3)
        if median_seq is None or median_seq in all_sequences:
            continue
        
        cost = (calc_dna_dist_cached(median_seq, s1) + 
                calc_dna_dist_cached(median_seq, s2) + 
                calc_dna_dist_cached(median_seq, s3))
        candidates.append((median_seq, cost, triplet))

    if not candidates:
        print(f"  No new medians found. Stopping.")
        break

    candidates.sort(key=lambda x: x[1])
    top_candidates = candidates[:5]

    added_this_iter = 0
    for median_seq, cost, triplet in top_candidates:
        if medians_added >= MAX_MEDIANS:
            break
        if median_seq in all_sequences:
            continue
            
        new_idx = len(all_seq_list)
        all_seq_list.append(median_seq)
        all_sequences.add(median_seq)
        median_flags.append(True)
        
        u, v, w = triplet
        for neighbor in [u, v, w]:
            adj[neighbor].add(new_idx)
            adj[new_idx].add(neighbor)
        
        medians_added += 1
        added_this_iter += 1
        changed = True

    if added_this_iter == 0:
        break

print(f"Added {medians_added} median vectors in {iteration} iterations.")

# ==========================================
# 6.5. ASSIGN HAPLOGROUPS TO MEDIAN VECTORS
# ==========================================
print("Assigning haplogroups to median vectors...")
median_haplogroups = {}
for idx in range(len(all_seq_list)):
    if not median_flags[idx]:
        continue
    neighbor_hgs = defaultdict(int)
    for neighbor_idx in adj.get(idx, []):
        if not median_flags[neighbor_idx]:
            neighbor_seq = all_seq_list[neighbor_idx]
            if neighbor_seq in hap_to_haplogroup:
                neighbor_hgs[hap_to_haplogroup[neighbor_seq]] += 1
    if neighbor_hgs:
        median_haplogroups[idx] = max(neighbor_hgs, key=neighbor_hgs.get)

print(f"Assigned haplogroups to {len(median_haplogroups)} median vectors.")

# ==========================================
# 7. GRAPH CONSTRUCTION (NETWORKX)
# ==========================================
print("Constructing NetworkX graph...")
COLORS = {
    'NAmer': '#1f77b4', 'SAmer': '#d62728', 'Eur': '#8c564b',
    'Asia': '#f1c40f', 'Austral': '#2ca02c', 'Unknown': '#cccccc'
}

def make_pie_svg(proportions, size=64):
    total = sum(proportions.values())
    cx, cy, r = size/2, size/2, size/2 - 2
    if total == 0:
        svg = f'<svg xmlns="http://www.w3.org/2000/svg" width="{size}" height="{size}">' \
              f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="#ffffff" stroke="#000000" ' \
              f'stroke-width="2" stroke-dasharray="3,2"/></svg>'
        return f"data:image/svg+xml;base64,{base64.b64encode(svg.encode()).decode()}"

    paths = []
    start_angle = 0
    for region, count in proportions.items():
        frac = count / total
        end_angle = start_angle + frac * 360
        color = COLORS.get(region, '#cccccc')
         
        if frac >= 0.999:
            paths.append(f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="{color}"/>')
            break
            
        x1 = cx + r * math.cos(math.radians(start_angle - 90))
        y1 = cy + r * math.sin(math.radians(start_angle - 90))
        x2 = cx + r * math.cos(math.radians(end_angle - 90))
        y2 = cy + r * math.sin(math.radians(end_angle - 90))
        large_arc = 1 if frac > 0.5 else 0
        
        paths.append(f'<path d="M {cx} {cy} L {x1} {y1} A {r} {r} 0 {large_arc} 1 {x2} {y2} Z" fill="{color}"/>')
        start_angle = end_angle

    paths.append(f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="#222" stroke-width="2"/>')
    svg_content = f'<svg xmlns="http://www.w3.org/2000/svg" width="{size}" height="{size}">' + ''.join(paths) + '</svg>'
    return f"data:image/svg+xml;base64,{base64.b64encode(svg_content.encode()).decode()}"

G = nx.Graph()
median_counter = 0

# Собираем маппинг node_id -> haplogroup для JS
node_to_haplogroup = {}

for idx, seq in enumerate(all_seq_list):
    if median_flags[idx] is None:
        continue
    is_median = median_flags[idx] == True

    if is_median:
        median_counter += 1
        node_id = f"mv_{median_counter:02d}"
        node_label = f"mv{median_counter}"
        total_traits = {}
        freq = 0
        hg = median_haplogroups.get(idx, None)
    else:
        hap_id = f"Hap_{idx+1:03d}"
        node_id = hap_id
        ids = haplotypes[seq]
        freq = len(ids)
        node_label = str(freq)
        hg = hap_to_haplogroup.get(seq, None)
        
        total_traits = defaultdict(int)
        for name in ids:
            if name in seq_traits:
                for region, count in seq_traits[name].items():
                    total_traits[region] += count

    # Сохраняем для JS
    if hg:
        node_to_haplogroup[node_id] = hg

    pie_img = make_pie_svg(total_traits, size=64)
    
    # Цвет рамки по гаплогруппе
    border_color = hg_colors.get(hg, '#999999') if hg else '#999999'
    
    if is_median:
        hg_text = f"Haplogroup: {hg} (inferred)\n" if hg else ""
        tooltip = (
            f"MEDIAN VECTOR: {node_id}\n"
            f"{hg_text}"
            f"(Not detected in the sample,\n"
            f"presumed ancestral haplotype)"
        )
        node_size = 10
        font_size = 12
    else:
        region_lines = [f"* {r}: {c}" for r, c in total_traits.items()]
        region_text = "\n".join(region_lines) if region_lines else "* No data"
        
        seq_lines = [f"* {sid}" for sid in haplotypes[seq][:15]]
        if len(haplotypes[seq]) > 15:
            seq_lines.append(f"...and {len(haplotypes[seq]) - 15} more")
        seq_text = "\n".join(seq_lines)
        
        hg_text = f"Haplogroup: {hg}\n" if hg else "Haplogroup: Unknown\n"
        
        tooltip = (
            f"Haplotype: {node_id}\n"
            f"Frequency: {freq} sequences\n"
            f"{hg_text}"
            f"{'-' * 25}\n"
            f"Regions:\n{region_text}\n"
            f"{'-' * 25}\n"
            f"Sequences:\n{seq_text}"
        )

        node_size = 20 + (freq * 2.0)
        font_size = 24

    G.add_node(
        node_id,
        label=" ",
        title=tooltip,
        shape='circularImage',
        image=pie_img,
        size=node_size,
        borderWidth=4 if not is_median else 3,
        borderColor=border_color,
        font={'size': font_size, 'color': '#000000', 'strokeWidth': 0, 'face': 'Arial'}
    )

# ==========================================
# 8. EDGES CONSTRUCTION
# ==========================================
print(f"Adding edges (TCS <= {TCS_LIMIT} mutations)...")

MAX_CONNECTIONS_PER_NODE = 4

def get_node_id(idx):
    if median_flags[idx]:
        median_num = sum(1 for k in range(idx + 1) if median_flags[k])
        return f"mv_{median_num:02d}"
    else:
        return f"Hap_{idx+1:03d}"

n_total = len(all_seq_list)

potential_edges = []
for i in range(n_total):
    if median_flags[i] is None: continue
    for j in range(i + 1, n_total):
        if median_flags[j] is None: continue
        
        d = calc_dna_dist_cached(all_seq_list[i], all_seq_list[j])
        if 0 < d <= TCS_LIMIT:
            potential_edges.append((d, i, j))

potential_edges.sort(key=lambda x: x[0])
node_degrees = {i: 0 for i in range(n_total)}
tcs_edges_count = 0

drawn_tcs_edges = []
for d, i, j in potential_edges:
    if node_degrees[i] < MAX_CONNECTIONS_PER_NODE and node_degrees[j] < MAX_CONNECTIONS_PER_NODE:
        id_i = get_node_id(i)
        id_j = get_node_id(j)
        
        edge_width = max(1, 10.0 - d * 1.0)
        G.add_edge(
            id_i, id_j,
            title=f"TCS: {d} mutations",
            color={'color': '#b0b0b0', 'highlight': '#333333'},
            value=edge_width,
            width=edge_width,
            length=100,          # ← ДОБАВИЛИ: длина пружинки
            smooth={'type': 'continuous', 'roundness': 0.1}
        )
        node_degrees[i] += 1
        node_degrees[j] += 1
        drawn_tcs_edges.append((i, j))
        tcs_edges_count += 1

print(f"Created {tcs_edges_count} cleaned TCS connections.")

print("Searching for VISUAL components...")
def find_visual_components(n_nodes, edges):
    adj_local = defaultdict(set)
    for i, j in edges:
        adj_local[i].add(j)
        adj_local[j].add(i)
        
    visited = set()
    components = []
    for i in range(n_nodes):
        if median_flags[i] is None or i in visited:
            continue
        
        comp = set()
        queue = [i]
        visited.add(i)
        
        while queue:
            curr = queue.pop(0)
            comp.add(curr)
            for nxt in adj_local[curr]:
                if nxt not in visited:
                    visited.add(nxt)
                    queue.append(nxt)
        components.append(comp)
    return components

components = find_visual_components(n_total, drawn_tcs_edges)
print(f"Found {len(components)} VISUAL components.")

print("Building MST between VISUAL clusters...")
class UnionFind:
    def __init__(self, n):
        self.parent = list(range(n))
    def find(self, i):
        if self.parent[i] == i: return i
        self.parent[i] = self.find(self.parent[i])
        return self.parent[i]
    def union(self, i, j):
        root_i = self.find(i)
        root_j = self.find(j)
        if root_i != root_j:
            self.parent[root_i] = root_j
            return True
        return False

component_edges = []
K = len(components)
for i in range(K):
    for j in range(i + 1, K):
        min_dist = float('inf')
        best_pair = None
        for idx_a in components[i]:
            for idx_b in components[j]:
                d = calc_dna_dist_cached(all_seq_list[idx_a], all_seq_list[idx_b])
                if d < min_dist:
                    min_dist = d
                    best_pair = (idx_a, idx_b)
        if best_pair:
            component_edges.append((min_dist, i, j, best_pair[0], best_pair[1]))

component_edges.sort(key=lambda x: x[0])
uf = UnionFind(K)
mst_bridges = []
for dist, c_i, c_j, idx_a, idx_b in component_edges:
    if uf.union(c_i, c_j):
        mst_bridges.append((idx_a, idx_b, dist))
        if len(mst_bridges) == K - 1:
            break

bridge_edges_count = 0
for idx_a, idx_b, dist in mst_bridges:
    id_a = get_node_id(idx_a)
    id_b = get_node_id(idx_b)
    G.add_edge(
        id_a, id_b,
        title=f"MST bridge: {dist} mutations\n(minimal connection between clusters)",
        color={'color': '#ff7f0e', 'highlight': '#d62728'},
        value=5.0,
        width=5.0,
        length=400,              # ← ДОБАВИЛИ: мосты длиннее, чтобы кластеры были дальше
        dashes=True,
        smooth={'type': 'curvedCW', 'roundness': 0.4}
    )
    bridge_edges_count += 1

print(f"Added {bridge_edges_count} MST bridges.")

# ==========================================
# 9. PYVIS CONVERSION & PHYSICS
# ==========================================
print("Converting to PyVis and applying physics...")
net = Network(height="100vh", width="100%", bgcolor="#ffffff", font_color="#000000", directed=False)

net.from_nx(G, show_edge_weights=False)

net.barnes_hut(
    gravity=-100000,          
    central_gravity=0.0001,    
    spring_length=10000,      
    spring_strength=0.000001,    
    damping=0.20             
)

net.set_options("""
{
  "edges": {
    "scaling": {
      "min": 1,
      "max": 10
    },
    "smooth": {
      "type": "continuous",
      "roundness": 0.1
    }
  },
  "interaction": {
    "dragNodes": true,
    "dragView": true,
    "zoomView": true,
    "hover": true
  }
}
""")

# ==========================================
# 10. HTML GENERATION + CONVEX HULL OVERLAY
# ==========================================
print("Generating final HTML with haplogroup hulls...")

# Готовим данные для JS
hg_map_json = json.dumps(node_to_haplogroup)
hg_colors_json = json.dumps(hg_colors)
hg_names_json = json.dumps(unique_haplogroups)

# JavaScript для рисования convex hull поверх сети
hull_js = f"""
<script>
// === HAPLOGROUP CONVEX HULL OVERLAY ===
const haplogroupMap = {hg_map_json};
const haplogroupColors = {hg_colors_json};
const haplogroupNames = {hg_names_json};

// Andrew's monotone chain convex hull algorithm
function convexHull(points) {{
    if (points.length < 3) return points;
    const pts = points.slice().sort((a, b) => a[0] - b[0] || a[1] - b[1]);
    const cross = (O, A, B) => (A[0]-O[0])*(B[1]-O[1]) - (A[1]-O[1])*(B[0]-O[0]);
    const lower = [];
    for (const p of pts) {{
        while (lower.length >= 2 && cross(lower[lower.length-2], lower[lower.length-1], p) <= 0) lower.pop();
        lower.push(p);
    }}
    const upper = [];
    for (let i = pts.length - 1; i >= 0; i--) {{
        const p = pts[i];
        while (upper.length >= 2 && cross(upper[upper.length-2], upper[upper.length-1], p) <= 0) upper.pop();
        upper.push(p);
    }}
    return lower.slice(0, -1).concat(upper.slice(0, -1));
}}

function drawHull(ctx, points, color, label, padding) {{
    if (points.length === 0) return;
    
    if (points.length === 1) {{
        // Один узел — рисуем круг
        ctx.beginPath();
        ctx.arc(points[0][0], points[0][1], padding, 0, 2 * Math.PI);
        ctx.fillStyle = color + '25';
        ctx.fill();
        ctx.strokeStyle = color + '25';
        ctx.lineWidth = 3;
        ctx.setLineDash([10, 6]);
        ctx.stroke();
        ctx.setLineDash([]);
        ctx.fillStyle = '#000000';
        ctx.font = 'bold 96px Arial';
        ctx.textAlign = 'center';
        ctx.textBaseline = 'middle';
        ctx.fillText(label, points[0][0], points[0][1] - padding - 10);
        return;
    }}
    
    if (points.length === 2) {{
        // Два узла — рисуем "капсулу"
        ctx.beginPath();
        ctx.moveTo(points[0][0], points[0][1]);
        ctx.lineTo(points[1][0], points[1][1]);
        ctx.strokeStyle = color + '25';
        ctx.lineWidth = padding * 2;
        ctx.lineCap = 'round';
        ctx.stroke();
        ctx.strokeStyle = color + '25';
        ctx.lineWidth = 3;
        ctx.setLineDash([10, 6]);
        ctx.stroke();
        ctx.setLineDash([]);
        const cx = (points[0][0] + points[1][0]) / 2;
        const cy = (points[0][1] + points[1][1]) / 2;
        ctx.fillStyle = '#000000';
        ctx.font = 'bold 96px Arial';
        ctx.textAlign = 'center';
        ctx.textBaseline = 'middle';
        ctx.fillText(label, cx, cy - padding - 10);
        return;
    }}
    
    // 3+ узлов — convex hull
    const hull = convexHull(points);
    
    // Центроид
    let cx = 0, cy = 0;
    for (const p of hull) {{ cx += p[0]; cy += p[1]; }}
    cx /= hull.length; cy /= hull.length;
    
    // Расширяем hull от центроида на padding
    const expanded = hull.map(p => {{
        const dx = p[0] - cx, dy = p[1] - cy;
        const dist = Math.sqrt(dx*dx + dy*dy) || 1;
        return [cx + dx * (1 + padding / dist), cy + dy * (1 + padding / dist)];
    }});
    
    // Рисуем заполнение
    ctx.beginPath();
    ctx.moveTo(expanded[0][0], expanded[0][1]);
    for (let i = 1; i < expanded.length; i++) {{
        ctx.lineTo(expanded[i][0], expanded[i][1]);
    }}
    ctx.closePath();
    ctx.fillStyle = color + '25';
    ctx.fill();
    
    // Рисуем пунктирную рамку
    ctx.strokeStyle = color + '25';
    ctx.lineWidth = 3;
    ctx.lineJoin = 'round';
    ctx.setLineDash([10, 6]);
    ctx.stroke();
    ctx.setLineDash([]);
    
    // Подпись гаплогруппы в центроиде
    ctx.fillStyle = '#000000';
    ctx.font = 'bold 96px Arial';
    ctx.textAlign = 'center';
    ctx.textBaseline = 'middle';
    ctx.fillText(label, cx, cy);
}}

// Рисуем hull при каждом кадре (автоматически масштабируется при зуме/перетаскивании)
network.on("afterDrawing", function(ctx) {{
    const positions = network.getPositions();
    
    // Группируем узлы по гаплогруппам
    const groups = {{}};
    for (const [nodeId, hg] of Object.entries(haplogroupMap)) {{
        if (positions[nodeId]) {{
            if (!groups[hg]) groups[hg] = [];
            groups[hg].push([positions[nodeId].x, positions[nodeId].y]);
        }}
    }}
    
    // Рисуем hull для каждой гаплогруппы
    for (const [hg, points] of Object.entries(groups)) {{
        const color = haplogroupColors[hg] || '#999999';
        drawHull(ctx, points, color, hg, 80);
    }}
}});
</script>
"""

# Легенда гаплогрупп (левый верхний угол)
hg_legend_items = ""
for hg in unique_haplogroups:
    color = hg_colors[hg]
    hg_legend_items += f'<span style="color:{color}; font-size: 18px; vertical-align: middle;">&#9632;</span> <span style="vertical-align: middle;">{hg}</span><br>\n'

hg_legend_html = f"""
<div style="position: fixed; top: 20px; left: 20px; background: rgba(255,255,255,0.97); 
border: 1px solid #ccc; padding: 18px 22px; border-radius: 10px; font-family: Arial, sans-serif; 
box-shadow: 2px 4px 12px rgba(0,0,0,0.2); z-index: 1000; font-size: 18px; color: #222; min-width: 200px; line-height: 1.4;">
    <b style="font-size: 22px; display: block; margin-bottom: 10px;">Haplogroups:</b>
    <div style="line-height: 1.6;">
        {hg_legend_items}
    </div>
    <hr style="margin: 10px 0; border: 0; border-top: 1px solid #ddd;">
    <div style="font-size: 13px; color: #888;">
        Colored border = haplogroup<br>
        Shaded area = clade outline
    </div>
</div>
"""

# Легенда регионов (правый верхний угол)
legend_html = f"""
<div style="position: fixed; top: 20px; right: 20px; background: rgba(255,255,255,0.97); 
border: 1px solid #ccc; padding: 18px 22px; border-radius: 10px; font-family: Arial, sans-serif; 
box-shadow: 2px 4px 12px rgba(0,0,0,0.2); z-index: 1000; font-size: 18px; color: #222; min-width: 260px; line-height: 1.4;">
    
    <b style="font-size: 22px; display: block; margin-bottom: 10px; line-height: 1.0;">Regions:</b>
    
    <div style="line-height: 0.75; margin-bottom: 4px;">
        <span style="color:#1f77b4; font-size: 48px; vertical-align: middle;">&#9679;</span> 
        <span style="vertical-align: middle;">North America</span><br>
        <span style="color:#d62728; font-size: 48px; vertical-align: middle;">&#9679;</span> 
        <span style="vertical-align: middle;">South America</span><br>
        <span style="color:#8c564b; font-size: 48px; vertical-align: middle;">&#9679;</span> 
        <span style="vertical-align: middle;">Europe</span><br>
        <span style="color:#f1c40f; font-size: 48px; vertical-align: middle;">&#9679;</span> 
        <span style="vertical-align: middle;">Asia</span><br>
        <span style="color:#2ca02c; font-size: 48px; vertical-align: middle;">&#9679;</span> 
        <span style="vertical-align: middle;">Australia</span>
    </div>

    <hr style="margin: 12px 0; border: 0; border-top: 2px solid #ddd;">
    
    <div style="font-size: 15px; line-height: 1.5;">
        <b>Algorithm:</b> Median-Joining<br>
        <b>Threshold:</b> TCS {int(TCS_CONFIDENCE*100)}%<br>
        <b>Parsimony limit:</b> {TCS_LIMIT} mut.<br>
        <b>Unique haplotypes:</b> {len(unique_haps)}<br>
        <b>Median vectors:</b> {median_counter}<br>
    </div>

    <hr style="margin: 12px 0; border: 0; border-top: 1px solid #ddd;">
    
    <div style="font-size: 15px; line-height: 1.5;">
        <b>Size</b> = frequency<br>
        <b>Line width</b> = mutations (thicker = fewer)<br>
        <b style="color:#ff7f0e;">┄ Dashed orange</b> = MST bridge<br>
        <b style="color:#666;">&#9711; Dashed circle</b> = median vector<br>
        <i style="font-size:13px; color:#888;">(ancestral haplotype)</i>
    </div>
</div>
"""

controls_html = """
<div id="physics-panel" style="position: fixed; bottom: 20px; right: 20px; background: rgba(255,255,255,0.97); 
border: 1px solid #ccc; padding: 16px 20px; border-radius: 10px; font-family: Arial, sans-serif; 
box-shadow: 2px 4px 12px rgba(0,0,0,0.2); z-index: 1000; font-size: 14px; color: #222; min-width: 320px;">
    
    <b style="font-size: 16px; display: block; margin-bottom: 10px;">&#9881; Physics Controls</b>
    
    <label style="display:block; margin-bottom:8px;">
        Spring Length: <span id="sl-val">400</span>
        <input type="range" id="sl-slider" min="100" max="3000" step="50" value="400" 
               style="width:100%; margin-top:2px;">
    </label>
    
    <label style="display:block; margin-bottom:8px;">
        Spring Stiffness: <span id="sc-val">5</span>
        <input type="range" id="sc-slider" min="1" max="50" step="1" value="5" 
               style="width:100%; margin-top:2px;">
    </label>
    
    <label style="display:block; margin-bottom:8px;">
        Repulsion: <span id="gr-val">50000</span>
        <input type="range" id="gr-slider" min="5000" max="300000" step="5000" value="50000" 
               style="width:100%; margin-top:2px;">
    </label>
    
    <label style="display:block; margin-bottom:8px;">
        Central Gravity: <span id="cg-val">10</span>
        <input type="range" id="cg-slider" min="1" max="100" step="1" value="10" 
               style="width:100%; margin-top:2px;">
    </label>
    
    <div style="margin-top:10px; display:flex; gap:8px;">
        <button onclick="network.stabilize()" 
                style="flex:1; padding:6px 12px; border:1px solid #aaa; border-radius:6px; 
                       background:#f0f0f0; cursor:pointer; font-size:13px;">
            &#8635; Re-stabilize
        </button>
        <button onclick="togglePhysics(this)" 
                style="flex:1; padding:6px 12px; border:1px solid #aaa; border-radius:6px; 
                       background:#f0f0f0; cursor:pointer; font-size:13px;">
            &#10074;&#10074; Stop Physics
        </button>
    </div>
</div>

<script>
let physicsRunning = true;

function togglePhysics(btn) {
    if (physicsRunning) {
        network.stopSimulation();
        btn.innerHTML = '&#9654; Start Physics';
        physicsRunning = false;
    } else {
        network.startSimulation();
        btn.innerHTML = '&#10074;&#10074; Stop Physics';
        physicsRunning = true;
    }
}

document.getElementById('sl-slider').addEventListener('input', function() {
    document.getElementById('sl-val').textContent = this.value;
    network.setOptions({ physics: { barnesHut: { springLength: parseInt(this.value) } } });
});

document.getElementById('sc-slider').addEventListener('input', function() {
    document.getElementById('sc-val').textContent = this.value;
    network.setOptions({ physics: { barnesHut: { springConstant: parseInt(this.value) / 10000 } } });
});

document.getElementById('gr-slider').addEventListener('input', function() {
    document.getElementById('gr-val').textContent = this.value;
    network.setOptions({ physics: { barnesHut: { gravitationalConstant: -parseInt(this.value) } } });
});

document.getElementById('cg-slider').addEventListener('input', function() {
    document.getElementById('cg-val').textContent = this.value;
    network.setOptions({ physics: { barnesHut: { centralGravity: parseInt(this.value) / 10000 } } });
});
</script>
"""


html_content = net.generate_html()
html_content = html_content.replace("</body>", hull_js + "\n" + controls_html + "\n" + hg_legend_html + "\n" + legend_html + "\n</body>")

with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
    f.write(html_content)

print(f"Done! File '{OUTPUT_HTML}' is ready.")
print(f"Summary: {len(unique_haps)} real haplotypes + {median_counter} median vectors.")
print(f"Haplogroups: {len(unique_haplogroups)} ({', '.join(unique_haplogroups)})")